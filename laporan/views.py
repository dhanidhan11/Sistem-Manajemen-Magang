from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db import models
from peserta.models import Peserta
from absensi.models import Absensi
from logbook.models import Logbook
from penilaian.models import Penilaian
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# LAPORAN PESERTA
# ============================================================

@login_required(login_url='/')
def laporan_peserta(request):
    """Laporan untuk peserta magang"""
    
    try:
        peserta = Peserta.objects.get(user=request.user)
    except Peserta.DoesNotExist:
        messages.error(request, 'Anda tidak terdaftar sebagai peserta magang!')
        return redirect('/dashboard/peserta/')
    
    # Hitung data
    total_absensi = Absensi.objects.filter(peserta=peserta).count()
    total_logbook = Logbook.objects.filter(peserta=peserta).count()
    
    # Hitung hari magang berjalan
    if peserta.tanggal_mulai:
        today = date.today()
        hari_ke = (today - peserta.tanggal_mulai).days + 1
        if hari_ke < 1:
            hari_ke = 0
    else:
        hari_ke = 0
    
    # Ambil penilaian
    penilaian = Penilaian.objects.filter(peserta=peserta).first()
    
    # Rekap status absensi
    status_absensi = Absensi.objects.filter(
        peserta=peserta
    ).values('status').annotate(
        total=models.Count('id')
    )
    
    # Tambahkan display name untuk status
    for item in status_absensi:
        item['get_status_display'] = dict(Absensi.STATUS_CHOICES).get(item['status'], item['status'])
    
    # Riwayat absensi
    riwayat_absensi = Absensi.objects.filter(
        peserta=peserta
    ).order_by('-tanggal')[:30]
    
    # Logbook terakhir
    logbooks = Logbook.objects.filter(
        peserta=peserta
    ).order_by('-tanggal')[:10]
    
    context = {
        'peserta': peserta,
        'total_absensi': total_absensi,
        'total_logbook': total_logbook,
        'hari_ke': hari_ke,
        'penilaian': penilaian,
        'status_absensi': status_absensi,
        'riwayat_absensi': riwayat_absensi,
        'logbooks': logbooks,
        'today': date.today(),
    }
    return render(request, 'laporan_peserta.html', context)


# ============================================================
# LAPORAN ADMIN
# ============================================================

@login_required(login_url='/')
def laporan_admin(request):
    """Laporan untuk admin (semua peserta)"""
    
    if not request.user.groups.filter(name='Admin').exists():
        messages.error(request, 'Anda tidak memiliki akses!')
        return redirect('/dashboard/admin/')
    
    total_peserta = Peserta.objects.count()
    total_peserta_aktif = Peserta.objects.filter(status='aktif').count()
    total_absensi_hari_ini = Absensi.objects.filter(tanggal=date.today()).count()
    
    # Rekap per bidang
    bidang_data = Peserta.objects.values('bidang_penempatan').annotate(
        total=models.Count('id')
    )
    
    # Rekap per mentor
    mentor_data = Peserta.objects.values('mentor__first_name', 'mentor__last_name').annotate(
        total=models.Count('id')
    )
    
    # Absensi 30 hari terakhir
    absensi_terakhir = Absensi.objects.all().order_by('-tanggal')[:30]
    
    context = {
        'total_peserta': total_peserta,
        'total_peserta_aktif': total_peserta_aktif,
        'total_absensi_hari_ini': total_absensi_hari_ini,
        'bidang_data': bidang_data,
        'mentor_data': mentor_data,
        'absensi_terakhir': absensi_terakhir,
        'today': date.today(),
    }
    return render(request, 'laporan_admin.html', context)


# ============================================================
# EXPORT EXCEL - PESERTA
# ============================================================

@login_required(login_url='/')
def export_excel_peserta(request):
    """Export laporan peserta ke Excel"""
    
    try:
        peserta = Peserta.objects.get(user=request.user)
    except Peserta.DoesNotExist:
        messages.error(request, 'Anda tidak terdaftar sebagai peserta magang!')
        return redirect('/dashboard/peserta/')
    
    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Magang"
    
    # Style
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # === HEADER ===
    ws.merge_cells('A1:F1')
    ws['A1'] = 'LAPORAN MAGANG - DISKOMINFOSANTIK'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # === DATA DIRI ===
    row = 3
    ws['A{}'.format(row)] = 'DATA DIRI'
    ws['A{}'.format(row)].font = bold_font
    ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    data_diri = [
        ['Nama', peserta.user.get_full_name()],
        ['NIM', peserta.nim],
        ['Institusi', peserta.institusi],
        ['Jurusan', peserta.jurusan],
        ['Bidang', peserta.bidang_penempatan],
        ['Mentor', peserta.mentor.get_full_name() if peserta.mentor else '-'],
        ['Tanggal Mulai', peserta.tanggal_mulai.strftime('%d/%m/%Y') if peserta.tanggal_mulai else '-'],
        ['Tanggal Selesai', peserta.tanggal_selesai.strftime('%d/%m/%Y') if peserta.tanggal_selesai else '-'],
        ['Status', peserta.status],
    ]
    
    for i, (label, value) in enumerate(data_diri):
        r = row + i + 1
        ws['A{}'.format(r)] = label
        ws['B{}'.format(r)] = value
        ws['A{}'.format(r)].font = bold_font
    
    row = row + len(data_diri) + 2
    
    # === STATISTIK ===
    ws['A{}'.format(row)] = 'STATISTIK'
    ws['A{}'.format(row)].font = bold_font
    ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    total_absensi = Absensi.objects.filter(peserta=peserta).count()
    total_logbook = Logbook.objects.filter(peserta=peserta).count()
    
    statistik = [
        ['Total Absensi', total_absensi],
        ['Total Logbook', total_logbook],
        ['Hari Ke-', (date.today() - peserta.tanggal_mulai).days + 1 if peserta.tanggal_mulai else 0],
    ]
    
    for i, (label, value) in enumerate(statistik):
        r = row + i + 1
        ws['A{}'.format(r)] = label
        ws['B{}'.format(r)] = value
        ws['A{}'.format(r)].font = bold_font
    
    row = row + len(statistik) + 2
    
    # === PENILAIAN ===
    penilaian = Penilaian.objects.filter(peserta=peserta).first()
    
    ws['A{}'.format(row)] = 'PENILAIAN'
    ws['A{}'.format(row)].font = bold_font
    ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    if penilaian:
        data_penilaian = [
            ['Kedisiplinan', penilaian.kedisiplinan],
            ['Kreativitas', penilaian.kreativitas],
            ['Komunikasi', penilaian.komunikasi],
            ['Kemampuan Teknis', penilaian.teknis],
            ['Presensi', penilaian.presensi],
            ['Presentasi', penilaian.presentasi],
            ['Sikap', penilaian.sikap],
            ['Total Nilai', penilaian.total_nilai],
        ]
        
        for i, (label, value) in enumerate(data_penilaian):
            r = row + i + 1
            ws['A{}'.format(r)] = label
            ws['B{}'.format(r)] = value
            ws['A{}'.format(r)].font = bold_font
    else:
        ws['A{}'.format(row + 1)] = 'Belum ada penilaian'
        data_penilaian = []
    
    row = row + (len(data_penilaian) if penilaian else 2) + 2
    
    # === RIWAYAT ABSENSI ===
    riwayat_absensi = Absensi.objects.filter(peserta=peserta).order_by('-tanggal')[:30]
    
    if riwayat_absensi:
        ws['A{}'.format(row)] = 'RIWAYAT ABSENSI (30 TERAKHIR)'
        ws['A{}'.format(row)].font = bold_font
        ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Header tabel
        headers = ['No', 'Tanggal', 'Check-in', 'Check-out', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data
        for i, absen in enumerate(riwayat_absensi, 1):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=absen.tanggal.strftime('%d/%m/%Y')).border = border
            ws.cell(row=r, column=3, value=absen.check_in.strftime('%H:%M') if absen.check_in else '-').border = border
            ws.cell(row=r, column=4, value=absen.check_out.strftime('%H:%M') if absen.check_out else '-').border = border
            ws.cell(row=r, column=5, value=absen.get_status_display()).border = border
        
        row = row + len(riwayat_absensi) + 2
    
    # === LOGBOOK ===
    logbooks = Logbook.objects.filter(peserta=peserta).order_by('-tanggal')[:10]
    
    if logbooks:
        ws['A{}'.format(row)] = 'LOGBOOK (10 TERAKHIR)'
        ws['A{}'.format(row)].font = bold_font
        ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Header
        headers = ['No', 'Tanggal', 'Kegiatan', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data
        for i, log in enumerate(logbooks, 1):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=log.tanggal.strftime('%d/%m/%Y')).border = border
            ws.cell(row=r, column=3, value=log.kegiatan[:100] + '...' if len(log.kegiatan) > 100 else log.kegiatan).border = border
            ws.cell(row=r, column=4, value=log.get_status_display()).border = border
    
    # === AUTO WIDTH ===
    for col in range(1, 7):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # === RESPONSE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Laporan_Magang_{peserta.user.get_full_name()}_{date.today().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================
# EXPORT EXCEL - ADMIN
# ============================================================

@login_required(login_url='/')
def export_excel_admin(request):
    """Export laporan admin ke Excel"""
    
    if not request.user.groups.filter(name='Admin').exists():
        messages.error(request, 'Anda tidak memiliki akses!')
        return redirect('/dashboard/admin/')
    
    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Admin"
    
    # Style
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # === HEADER ===
    ws.merge_cells('A1:F1')
    ws['A1'] = 'LAPORAN ADMIN - DISKOMINFOSANTIK'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # === STATISTIK ===
    row = 3
    ws['A{}'.format(row)] = 'STATISTIK'
    ws['A{}'.format(row)].font = bold_font
    ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    statistik = [
        ['Total Peserta', Peserta.objects.count()],
        ['Peserta Aktif', Peserta.objects.filter(status='aktif').count()],
        ['Absensi Hari Ini', Absensi.objects.filter(tanggal=date.today()).count()],
    ]
    
    for i, (label, value) in enumerate(statistik):
        r = row + i + 1
        ws['A{}'.format(r)] = label
        ws['B{}'.format(r)] = value
        ws['A{}'.format(r)].font = bold_font
    
    row = row + len(statistik) + 2
    
    # === PER BIDANG ===
    bidang_data = Peserta.objects.values('bidang_penempatan').annotate(
        total=models.Count('id')
    )
    
    if bidang_data:
        ws['A{}'.format(row)] = 'PESERTA PER BIDANG'
        ws['A{}'.format(row)].font = bold_font
        ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        headers = ['No', 'Bidang', 'Jumlah']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        for i, item in enumerate(bidang_data, 1):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=item['bidang_penempatan']).border = border
            ws.cell(row=r, column=3, value=item['total']).border = border
        
        row = row + len(bidang_data) + 2
    
    # === PER MENTOR ===
    mentor_data = Peserta.objects.values('mentor__first_name', 'mentor__last_name').annotate(
        total=models.Count('id')
    )
    
    if mentor_data:
        ws['A{}'.format(row)] = 'PESERTA PER MENTOR'
        ws['A{}'.format(row)].font = bold_font
        ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        headers = ['No', 'Mentor', 'Jumlah']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        for i, item in enumerate(mentor_data, 1):
            r = row + 1 + i
            mentor_name = f"{item['mentor__first_name']} {item['mentor__last_name']}".strip() or '-'
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=mentor_name).border = border
            ws.cell(row=r, column=3, value=item['total']).border = border
        
        row = row + len(mentor_data) + 2
    
    # === ABSENSI TERAKHIR ===
    absensi_terakhir = Absensi.objects.all().order_by('-tanggal')[:50]
    
    if absensi_terakhir:
        ws['A{}'.format(row)] = 'ABSENSI TERAKHIR (50)'
        ws['A{}'.format(row)].font = bold_font
        ws['A{}'.format(row)].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        headers = ['No', 'Peserta', 'Tanggal', 'Check-in', 'Check-out', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        for i, absen in enumerate(absensi_terakhir, 1):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=absen.peserta.user.get_full_name()).border = border
            ws.cell(row=r, column=3, value=absen.tanggal.strftime('%d/%m/%Y')).border = border
            ws.cell(row=r, column=4, value=absen.check_in.strftime('%H:%M') if absen.check_in else '-').border = border
            ws.cell(row=r, column=5, value=absen.check_out.strftime('%H:%M') if absen.check_out else '-').border = border
            ws.cell(row=r, column=6, value=absen.get_status_display()).border = border
    
    # === AUTO WIDTH ===
    for col in range(1, 7):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # === RESPONSE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Laporan_Admin_{date.today().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response